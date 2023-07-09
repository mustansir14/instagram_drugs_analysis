from typing import List

import openpyxl
from sqlalchemy import select, text

from instagram_scraper import models
from instagram_scraper.db import DB


class ReportGenerator:
    def __init__(self) -> None:
        self.db = DB.create()

    def generate_report(self, festivals: List[models.Festival] | None = None):
        if festivals is None:
            festivals = self.db.scalars(select(models.Festival))

        if not festivals:
            raise Exception("No festivals found!!")

        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)

        for festival in festivals:
            sheet = workbook.create_sheet(festival.name)
            self.fill_cell(sheet, 1, 1, "Total")
            sheet.merge_cells("B1:C1")
            self.fill_cell(sheet, 1, 2, "Substances")

            substances_dict = {}
            for substance_type in self.db.scalars(select(models.SubstanceType)):
                substances_dict[substance_type] = self.db.scalars(
                    select(models.Substance)
                    .where(models.Substance.nickname_of_substance_id == None)
                    .where(models.Substance.substance_type == substance_type)
                )

            row = 2
            header_row = 2
            max_column = 0
            for substance_type, substances in substances_dict.items():
                substance_type_row = row
                substance_type_total = 0
                self.fill_cell(sheet, row, 2, substance_type.name)
                row += 1
                for substance in substances:
                    substance_total = self.count_and_fill(
                        sheet, row, 2, header_row, substance, festival
                    )
                    nickname_col = 4
                    for nickname in substance.nicknames:
                        substance_total += self.count_and_fill(
                            sheet, row, nickname_col, header_row, nickname, festival
                        )
                        nickname_col += 2

                    if nickname_col > max_column:
                        max_column = nickname_col

                    self.fill_cell(sheet, row, 1, substance_total)
                    substance_type_total += substance_total
                    row += 1

                self.fill_cell(sheet, substance_type_row, 1, substance_type_total)
                row += 1

            sheet.merge_cells(
                start_row=1, start_column=4, end_row=1, end_column=max_column - 1
            )
            self.fill_cell(sheet, 1, 4, "Nicknames")
            row += 1
            self.fill_cell(sheet, row, 1, "Total Posts Scraped:")
            self.fill_cell(
                sheet,
                row,
                2,
                self.db.execute(
                    text(
                        f"""select count(distinct p.id) from post p 
                    inner join hashtag_post hp on p.id = hp.post_id
                    inner join hashtag h on hp.hashtag_id = h.id
                    inner join festival f on h.festival_id = f.id
                    where f.id = {festival.id};"""
                    )
                ).first()[0],
            )

        workbook.save("report.xlsx")

    def count_and_fill(
        self,
        sheet,
        row: int,
        col: int,
        header_row: int,
        substance: models.Substance,
        festival: models.Festival,
    ) -> int:
        self.fill_cell(sheet, row, col, substance.name)
        self.fill_cell(sheet, header_row, col + 1, "Hits")
        count = self.db.execute(
            text(
                f"""select count(distinct p.id) from post p 
                    inner join substance_post sp on p.id = sp.post_id 
                    inner join substance s on s.id = sp.substance_id 
                    inner join hashtag_post hp on p.id = hp.post_id
                    inner join hashtag h on hp.hashtag_id = h.id
                    inner join festival f on h.festival_id = f.id
                    where s.id = {substance.id} and f.id = {festival.id};"""
            )
        ).first()[0]
        self.fill_cell(sheet, row, col + 1, count)
        return count

    def fill_cell(self, sheet, row: int, col: int, value) -> None:
        sheet.cell(row=row, column=col).value = value


if __name__ == "__main__":
    report_generator = ReportGenerator()
    report_generator.generate_report()
